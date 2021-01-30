# 合并excel表格多个工作区
import openpyxl

source_name = './test.xlsx'
target_name = './target.xlsx'
wb = openpyxl.load_workbook(filename=source_name)
print(wb.sheetnames)
activews = wb[wb.sheetnames[0]]


# for cell in wb[wb.sheetnames[0]].rows.__next__():
#     title.append(cell.value)
# savews.append(title)

for ws in wb:
    # 第一个工作表不做处理
    if ws.title == activews.title:
        continue

    # 删除头行
    ws.delete_rows(0)
    for row in ws.rows:
        activews.append([cell.value for cell in row])
    wb.remove(ws)

# 保存
wb.save(target_name)
